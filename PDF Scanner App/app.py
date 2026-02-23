import os
import json
import uuid
import sys
from datetime import datetime, timedelta
from pathlib import Path

from flask import Flask, render_template, request, jsonify, send_file
from PyPDF2 import PdfReader
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "Output"
DATA_DIR = BASE_DIR / "data"
HISTORY_FILE = DATA_DIR / "history.json"
EXCEL_FILE = DATA_DIR / "history.xlsx"

OUTPUT_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)

COST_PER_PAGE = 2000.0 / (700 * 30)  # ~0.095238 EUR
MAX_FILES = 20
MAX_FILE_SIZE_MB = 100

# ---------------------------------------------------------------------------
# History helpers
# ---------------------------------------------------------------------------

def load_history():
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"uploads": [], "settings": {"monthly_income": 2000, "daily_pages": 700, "days_per_month": 30}}


def save_history(data):
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    save_excel(data)


def save_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Upload History"
    ws.append(["ID", "Data", "Ora", "Fisier", "Pagini", "Cost (EUR)", "Marime (MB)"])
    for u in data["uploads"]:
        ts = u.get("timestamp", u["date"])
        date_part = ts[:10]
        time_part = ts[11:19] if len(ts) > 10 else ""
        size_mb = round(u.get("size_bytes", 0) / (1024 * 1024), 2)
        ws.append([u["id"][:8], date_part, time_part, u["filename"], u["pages"], round(u["cost"], 4), size_mb])

    ws_daily = wb.create_sheet("Sumar Zilnic")
    ws_daily.append(["Data", "Fisiere", "Pagini", "Cost (EUR)"])
    daily = {}
    for u in data["uploads"]:
        d = u["timestamp"][:10]
        if d not in daily:
            daily[d] = {"files": 0, "pages": 0, "cost": 0.0}
        daily[d]["files"] += 1
        daily[d]["pages"] += u["pages"]
        daily[d]["cost"] += u["cost"]
    for d in sorted(daily.keys(), reverse=True):
        ws_daily.append([d, daily[d]["files"], daily[d]["pages"], round(daily[d]["cost"], 4)])

    wb.save(str(EXCEL_FILE))


def count_pdf_pages(filepath):
    try:
        reader = PdfReader(str(filepath))
        return len(reader.pages)
    except Exception as e:
        print(f"Error reading {filepath}: {e}", file=sys.stderr)
        return 0

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/upload", methods=["POST"])
def upload_files():
    if "files" not in request.files:
        return jsonify({"error": "Nu s-au trimis fisiere"}), 400

    files = request.files.getlist("files")
    if len(files) > MAX_FILES:
        return jsonify({"error": f"Maxim {MAX_FILES} fisiere permise"}), 400

    history = load_history()
    existing_names = {u["filename"] for u in history["uploads"]}
    results = []
    now = datetime.now()

    for f in files:
        if not f.filename or not f.filename.lower().endswith(".pdf"):
            results.append({"filename": f.filename or "unknown", "error": "Nu este PDF", "pages": 0, "cost": 0})
            continue

        if f.filename in existing_names:
            results.append({"filename": f.filename, "error": "Duplicat - exista deja in istoric", "pages": 0, "cost": 0})
            continue

        file_id = uuid.uuid4().hex[:12]
        safe_name = f"{file_id}_{f.filename}"
        save_path = OUTPUT_DIR / safe_name
        f.save(str(save_path))

        size_bytes = save_path.stat().st_size
        if size_bytes > MAX_FILE_SIZE_MB * 1024 * 1024:
            save_path.unlink()
            results.append({"filename": f.filename, "error": f"Fisier prea mare (max {MAX_FILE_SIZE_MB}MB)", "pages": 0, "cost": 0})
            continue

        pages = count_pdf_pages(save_path)
        cost = round(pages * COST_PER_PAGE, 4)

        entry = {
            "id": file_id,
            "date": now.strftime("%Y-%m-%d"),
            "timestamp": now.isoformat(timespec="seconds"),
            "filename": f.filename,
            "saved_as": safe_name,
            "pages": pages,
            "cost": cost,
            "size_bytes": size_bytes,
        }
        history["uploads"].append(entry)
        results.append({"filename": f.filename, "pages": pages, "cost": cost, "size_mb": round(size_bytes / 1048576, 2), "id": file_id})

    save_history(history)
    total_pages = sum(r.get("pages", 0) for r in results)
    total_cost = round(sum(r.get("cost", 0) for r in results), 4)

    return jsonify({"results": results, "total_pages": total_pages, "total_cost": total_cost, "cost_per_page": round(COST_PER_PAGE, 6)})


@app.route("/api/history")
def get_history():
    history = load_history()
    date_from = request.args.get("from")
    date_to = request.args.get("to")
    search = request.args.get("search", "").lower()

    uploads = history["uploads"]

    if date_from:
        uploads = [u for u in uploads if u["timestamp"][:10] >= date_from]
    if date_to:
        uploads = [u for u in uploads if u["timestamp"][:10] <= date_to]
    if search:
        uploads = [u for u in uploads if search in u["filename"].lower()]

    uploads = sorted(uploads, key=lambda x: x["timestamp"], reverse=True)

    total_pages = sum(u["pages"] for u in uploads)
    total_cost = round(sum(u["cost"] for u in uploads), 4)
    total_files = len(uploads)

    return jsonify({
        "uploads": uploads,
        "total_pages": total_pages,
        "total_cost": total_cost,
        "total_files": total_files,
        "cost_per_page": round(COST_PER_PAGE, 6),
    })


@app.route("/api/daily-summary")
def daily_summary():
    history = load_history()
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    uploads = history["uploads"]
    if date_from:
        uploads = [u for u in uploads if u["timestamp"][:10] >= date_from]
    if date_to:
        uploads = [u for u in uploads if u["timestamp"][:10] <= date_to]

    daily = {}
    for u in uploads:
        d = u["timestamp"][:10]
        if d not in daily:
            daily[d] = {"date": d, "files": 0, "pages": 0, "cost": 0.0, "filenames": []}
        daily[d]["files"] += 1
        daily[d]["pages"] += u["pages"]
        daily[d]["cost"] += u["cost"]
        daily[d]["filenames"].append(u["filename"])

    for d in daily:
        daily[d]["cost"] = round(daily[d]["cost"], 4)

    result = sorted(daily.values(), key=lambda x: x["date"], reverse=True)
    return jsonify({"days": result})


@app.route("/api/monthly-summary")
def monthly_summary():
    history = load_history()
    monthly = {}
    for u in history["uploads"]:
        m = u["timestamp"][:7]  # YYYY-MM
        if m not in monthly:
            monthly[m] = {"month": m, "files": 0, "pages": 0, "cost": 0.0, "days_active": set()}
        monthly[m]["files"] += 1
        monthly[m]["pages"] += u["pages"]
        monthly[m]["cost"] += u["cost"]
        monthly[m]["days_active"].add(u["timestamp"][:10])

    result = []
    for m in sorted(monthly.keys(), reverse=True):
        entry = monthly[m]
        entry["cost"] = round(entry["cost"], 4)
        entry["days_active"] = len(entry["days_active"])
        result.append(entry)

    return jsonify({"months": result})


@app.route("/api/stats")
def get_stats():
    history = load_history()
    uploads = history["uploads"]
    today = datetime.now().strftime("%Y-%m-%d")

    today_uploads = [u for u in uploads if u["timestamp"][:10] == today]
    today_pages = sum(u["pages"] for u in today_uploads)
    today_cost = round(sum(u["cost"] for u in today_uploads), 4)
    today_files = len(today_uploads)

    total_pages = sum(u["pages"] for u in uploads)
    total_cost = round(sum(u["cost"] for u in uploads), 4)
    total_files = len(uploads)

    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    week_uploads = [u for u in uploads if u["timestamp"][:10] >= week_ago]
    week_pages = sum(u["pages"] for u in week_uploads)
    week_cost = round(sum(u["cost"] for u in week_uploads), 4)

    month_start = datetime.now().strftime("%Y-%m-01")
    month_uploads = [u for u in uploads if u["timestamp"][:10] >= month_start]
    month_pages = sum(u["pages"] for u in month_uploads)
    month_cost = round(sum(u["cost"] for u in month_uploads), 4)

    return jsonify({
        "today": {"files": today_files, "pages": today_pages, "cost": today_cost},
        "week": {"pages": week_pages, "cost": week_cost, "files": len(week_uploads)},
        "month": {"pages": month_pages, "cost": month_cost, "files": len(month_uploads)},
        "total": {"files": total_files, "pages": total_pages, "cost": total_cost},
        "cost_per_page": round(COST_PER_PAGE, 6),
    })


@app.route("/api/delete/<upload_id>", methods=["DELETE"])
def delete_upload(upload_id):
    history = load_history()
    entry = next((u for u in history["uploads"] if u["id"] == upload_id), None)
    if not entry:
        return jsonify({"error": "Nu s-a gasit"}), 404

    filepath = OUTPUT_DIR / entry.get("saved_as", "")
    if filepath.exists():
        filepath.unlink()

    history["uploads"] = [u for u in history["uploads"] if u["id"] != upload_id]
    save_history(history)
    return jsonify({"success": True})


@app.route("/api/delete-bulk", methods=["POST"])
def delete_bulk():
    data = request.get_json()
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "Niciun ID specificat"}), 400

    history = load_history()
    deleted = 0
    for uid in ids:
        entry = next((u for u in history["uploads"] if u["id"] == uid), None)
        if entry:
            filepath = OUTPUT_DIR / entry.get("saved_as", "")
            if filepath.exists():
                filepath.unlink()
            deleted += 1
    history["uploads"] = [u for u in history["uploads"] if u["id"] not in ids]
    save_history(history)
    return jsonify({"success": True, "deleted": deleted})


@app.route("/api/reset-period", methods=["POST"])
def reset_period():
    data = request.get_json()
    date_from = data.get("from")
    date_to = data.get("to")

    if not date_from or not date_to:
        return jsonify({"error": "Specifica perioada (from, to)"}), 400

    history = load_history()
    to_delete = [u for u in history["uploads"] if date_from <= u["timestamp"][:10] <= date_to]

    for entry in to_delete:
        filepath = OUTPUT_DIR / entry.get("saved_as", "")
        if filepath.exists():
            filepath.unlink()

    deleted = len(to_delete)
    history["uploads"] = [u for u in history["uploads"] if not (date_from <= u["timestamp"][:10] <= date_to)]
    save_history(history)
    return jsonify({"success": True, "deleted": deleted})


@app.route("/api/export-excel")
def export_excel():
    if not EXCEL_FILE.exists():
        history = load_history()
        save_excel(history)
    return send_file(str(EXCEL_FILE), as_attachment=True, download_name="upload_history.xlsx")


@app.route("/api/settings", methods=["GET", "POST"])
def settings():
    history = load_history()
    if request.method == "POST":
        data = request.get_json()
        history["settings"] = {
            "monthly_income": float(data.get("monthly_income", 2000)),
            "daily_pages": int(data.get("daily_pages", 700)),
            "days_per_month": int(data.get("days_per_month", 30)),
        }
        save_history(history)
        s = history["settings"]
        new_cost = s["monthly_income"] / (s["daily_pages"] * s["days_per_month"])
        return jsonify({"success": True, "cost_per_page": round(new_cost, 6)})
    return jsonify(history.get("settings", {"monthly_income": 2000, "daily_pages": 700, "days_per_month": 30}))


if __name__ == "__main__":
    print(f"  Output folder: {OUTPUT_DIR}")
    print(f"  Cost per page: {COST_PER_PAGE:.6f} EUR")
    print(f"  History: {HISTORY_FILE}")
    app.run(debug=True, port=5000, use_reloader=False)
